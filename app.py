from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, g, send_file, send_from_directory, Response
from flask_bcrypt import Bcrypt
import os
import sqlite3
from datetime import datetime
import csv
from io import StringIO, BytesIO
try:
    import openpyxl
except Exception:
    openpyxl = None
from werkzeug.utils import secure_filename
from weasyprint import HTML, CSS

# Imaging & codes
from PIL import Image, ImageDraw, ImageFont
from barcode import EAN13
from barcode.writer import ImageWriter
import qrcode
from math import ceil
import json
import zipfile

# Optional: human-friendly number to words
try:
    from num2words import num2words
except Exception:
    num2words = None

DB_PATH = os.path.join(os.path.dirname(__file__), 'crm.db')

app = Flask(__name__)
# Static folder for item images
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = os.urandom(24)  # For session management
bcrypt = Bcrypt(app)

# Configure JSON provider to handle sqlite3.Row objects globally
from flask.json.provider import DefaultJSONProvider

class SqliteRowJSONProvider(DefaultJSONProvider):
    def default(self, o):
        # Convert sqlite3.Row to dict for JSON serialization
        if isinstance(o, sqlite3.Row):
            return dict(o)
        return super().default(o)

app.json_provider_class = SqliteRowJSONProvider
app.json = app.json_provider_class(app)

# ----------------------
# Jinja filter: Amount in words (Indian numbering)
# ----------------------

def _inr_number_to_words(n: int) -> str:
    ones = ["zero", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine",
            "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen", "seventeen", "eighteen", "nineteen"]
    tens = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]

    def two_digits(num):
        if num < 20:
            return ones[num]
        t, o = divmod(num, 10)
        return (tens[t] + (" " + ones[o] if o else "")).strip()

    def three_digits(num):
        h, rem = divmod(num, 100)
        parts = []
        if h:
            parts.append(ones[h] + " hundred")
        if rem:
            parts.append(two_digits(rem))
        return " ".join(parts) if parts else "zero"

    if n == 0:
        return "zero"

    parts = []
    crore, rem = divmod(n, 10_000_000)
    if crore:
        parts.append(_inr_number_to_words(crore) + " crore")
    lakh, rem = divmod(rem, 100_000)
    if lakh:
        parts.append(_inr_number_to_words(lakh) + " lakh")
    thousand, rem = divmod(rem, 1000)
    if thousand:
        parts.append(_inr_number_to_words(thousand) + " thousand")
    if rem:
        parts.append(three_digits(rem))

    return " ".join(parts)


def inr_words(amount) -> str:
    try:
        n = round(float(amount or 0), 2)
    except Exception:
        n = 0.0
    rupees = int(n)
    paise = int(round((n - rupees) * 100))

    if num2words:
        # Use library if available (en_IN style, currency-like phrasing)
        try:
            words = num2words(rupees, lang='en_IN').replace('-', ' ') + " rupees"
            if paise:
                words += " and " + num2words(paise, lang='en_IN').replace('-', ' ') + " paise"
            return words + " only"
        except Exception:
            pass

    # Fallback to lightweight implementation
    words = _inr_number_to_words(rupees) + " rupees"
    if paise:
        words += " and " + _inr_number_to_words(paise) + " paise"
    words += " only"
    return words

# register filter for templates
app.jinja_env.filters['inr_words'] = inr_words

# --- Add Date format filter here ---
@app.template_filter('dateformat')
def dateformat(value, format='%d-%m-%Y'):
    """Convert YYYY-MM-DD (or ISO datetime string) to given format."""
    if not value:
        return ""
    try:
        # Try YYYY-MM-DD first
        return datetime.strptime(value, "%Y-%m-%d").strftime(format)
    except ValueError:
        try:
            # Try full ISO timestamp: YYYY-MM-DD HH:MM:SS
            return datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime(format)
        except Exception:
            return value  # fallback: return as-is
# --- end filter ---

# ----------------------
# Database helpers
# ----------------------

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    db = get_db()

    # Users
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
    )

    # Customers
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT,
            mobile TEXT,
            phone TEXT,
            address TEXT,
            city TEXT,
            pincode TEXT,
            state TEXT,
            country TEXT,
            notes TEXT,
            status TEXT DEFAULT 'active',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
    )

    # Leads
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT,
            mobile TEXT,
            phone TEXT,
            source TEXT,
            status TEXT DEFAULT 'new',
            notes TEXT,
            customer_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(customer_id) REFERENCES customers(id)
        );
        """
    )

    # Opportunities
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS opportunities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            amount REAL,
            stage TEXT DEFAULT 'prospecting',
            status TEXT DEFAULT 'open',
            customer_id INTEGER,
            lead_id INTEGER,
            notes TEXT,
            close_date TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(customer_id) REFERENCES customers(id),
            FOREIGN KEY(lead_id) REFERENCES leads(id)
        );
        """
    )

    # Inventory Items
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT UNIQUE,
            name TEXT NOT NULL,
            description TEXT,
            uom TEXT DEFAULT 'Nos',
            cost_price REAL DEFAULT 0,
            selling_price REAL DEFAULT 0,
            stock_qty REAL DEFAULT 0,
            reorder_level REAL DEFAULT 0,
            hsn_code TEXT,
            gst_rate REAL DEFAULT 0,
            discount REAL DEFAULT 0,
            supplier_name TEXT,
            brand TEXT,
            image_path TEXT,
            status TEXT DEFAULT 'active',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
    )



    # Suppliers
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            address TEXT,
            gstin TEXT,
            status TEXT DEFAULT 'active',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
    )

    # Migration-lite: add supplier columns if missing
    try:
        cols = db.execute('PRAGMA table_info(suppliers)').fetchall()
        names = {c['name'] for c in cols}
        if 'city' not in names:
            db.execute("ALTER TABLE suppliers ADD COLUMN city TEXT")
        if 'pincode' not in names:
            db.execute("ALTER TABLE suppliers ADD COLUMN pincode TEXT")
        if 'state' not in names:
            db.execute("ALTER TABLE suppliers ADD COLUMN state TEXT")
        if 'gst_type' not in names:
            db.execute("ALTER TABLE suppliers ADD COLUMN gst_type TEXT DEFAULT 'regular'")
        if 'contact_person' not in names:
            db.execute("ALTER TABLE suppliers ADD COLUMN contact_person TEXT")
        if 'mobile' not in names:
            db.execute("ALTER TABLE suppliers ADD COLUMN mobile TEXT")
    except Exception:
        pass

    # Ensure items table has required columns (migration-lite)
    try:
        cols = db.execute('PRAGMA table_info(items)').fetchall()
        names = {c['name'] for c in cols}
        # Core numeric columns used across the app
        if 'cost_price' not in names:
            db.execute("ALTER TABLE items ADD COLUMN cost_price REAL DEFAULT 0")
        if 'selling_price' not in names:
            db.execute("ALTER TABLE items ADD COLUMN selling_price REAL DEFAULT 0")
        if 'stock_qty' not in names:
            db.execute("ALTER TABLE items ADD COLUMN stock_qty REAL DEFAULT 0")
        if 'reorder_level' not in names:
            db.execute("ALTER TABLE items ADD COLUMN reorder_level REAL DEFAULT 0")
        # Other columns referenced by routes/APIs
        if 'uom' not in names:
            db.execute("ALTER TABLE items ADD COLUMN uom TEXT DEFAULT 'Nos'")
        if 'hsn_code' not in names:
            db.execute("ALTER TABLE items ADD COLUMN hsn_code TEXT")
        if 'gst_rate' not in names:
            db.execute("ALTER TABLE items ADD COLUMN gst_rate REAL DEFAULT 0")
        if 'discount' not in names:
            db.execute("ALTER TABLE items ADD COLUMN discount REAL DEFAULT 0")
        if 'supplier_id' not in names:
            db.execute("ALTER TABLE items ADD COLUMN supplier_id INTEGER")
        if 'supplier_name' not in names:
            db.execute("ALTER TABLE items ADD COLUMN supplier_name TEXT")
        if 'brand' not in names:
            db.execute("ALTER TABLE items ADD COLUMN brand TEXT")
        if 'image_path' not in names:
            db.execute("ALTER TABLE items ADD COLUMN image_path TEXT")
        if 'status' not in names:
            db.execute("ALTER TABLE items ADD COLUMN status TEXT DEFAULT 'active'")
        if 'ean13' not in names:
            db.execute("ALTER TABLE items ADD COLUMN ean13 TEXT")
    except Exception as e:
        try:
            print('Item table migration check error:', e)
        except Exception:
            pass

    # Stock Moves (audit trail)
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS stock_moves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL,
            move_type TEXT NOT NULL CHECK (move_type IN ('IN','OUT')),
            qty REAL NOT NULL,
            reference TEXT,
            reference_id INTEGER,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(item_id) REFERENCES items(id)
        );
        """
    )

    # Purchases Orders (basic - kept for compatibility)
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS purchases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER NOT NULL,
            date TEXT,
            bill_no TEXT,
            notes TEXT,
            total REAL DEFAULT 0,
            tax_total REAL DEFAULT 0,
            grand_total REAL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(supplier_id) REFERENCES suppliers(id)
        );
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS purchase_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_id INTEGER NOT NULL,
            item_id INTEGER NOT NULL,
            qty REAL NOT NULL,
            rate REAL NOT NULL,
            discount REAL DEFAULT 0,
            gst_rate REAL DEFAULT 0,
            line_total REAL DEFAULT 0,
            FOREIGN KEY(purchase_id) REFERENCES purchases(id),
            FOREIGN KEY(item_id) REFERENCES items(id)
        );
        """
    )

    # Payments (Customer receipts)
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_id INTEGER NOT NULL,
            date TEXT,
            mode TEXT,
            reference TEXT,
            amount REAL NOT NULL,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(invoice_id) REFERENCES sales_invoices(id)
        );
        """
    )

    # Supplier Payments (Payables against purchase invoices)
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS supplier_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_invoice_id INTEGER NOT NULL,
            date TEXT,
            mode TEXT,
            reference TEXT,
            amount REAL NOT NULL,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(purchase_invoice_id) REFERENCES purchase_invoices(id)
        );
        """
    )

# Sales Orders
    db.execute("""
        CREATE TABLE IF NOT EXISTS sales_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            so_no TEXT UNIQUE,
            customer_id INTEGER NOT NULL,
            date TEXT,
            expected_delivery_date TEXT,
            notes TEXT,
            total REAL DEFAULT 0,
            grand_total REAL DEFAULT 0,
            status TEXT DEFAULT 'draft',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(customer_id) REFERENCES customers(id)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS sales_order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sales_order_id INTEGER NOT NULL,
            item_id INTEGER NOT NULL,
            qty REAL NOT NULL,
            rate REAL NOT NULL,
            discount REAL DEFAULT 0,
            line_total REAL DEFAULT 0,
            FOREIGN KEY(sales_order_id) REFERENCES sales_orders(id),
            FOREIGN KEY(item_id) REFERENCES items(id)
        )
    """)

    # Sales Invoices
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS sales_invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            si_no TEXT UNIQUE,
            customer_id INTEGER NOT NULL,
            date TEXT,
            notes TEXT,
            total REAL DEFAULT 0,
            tax_total REAL DEFAULT 0,
            grand_total REAL DEFAULT 0,
            status TEXT DEFAULT 'draft',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(customer_id) REFERENCES customers(id)
        );
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS sales_invoice_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_id INTEGER NOT NULL,
            item_id INTEGER NOT NULL,
            qty REAL NOT NULL,
            rate REAL NOT NULL,
            discount REAL DEFAULT 0,
            gst_rate REAL DEFAULT 0,
            line_total REAL DEFAULT 0,
            FOREIGN KEY(invoice_id) REFERENCES sales_invoices(id),
            FOREIGN KEY(item_id) REFERENCES items(id)
        );
        """
    )

    # Purchase Invoices
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS purchase_invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pi_no TEXT UNIQUE,
            supplier_id INTEGER NOT NULL,
            date TEXT,
            bill_no TEXT,
            notes TEXT,
            total REAL DEFAULT 0,
            tax_total REAL DEFAULT 0,
            grand_total REAL DEFAULT 0,
            status TEXT DEFAULT 'draft',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(supplier_id) REFERENCES suppliers(id)
        );
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS purchase_invoice_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_id INTEGER NOT NULL,
            item_id INTEGER NOT NULL,
            qty REAL NOT NULL,
            rate REAL NOT NULL,
            discount REAL DEFAULT 0,
            gst_rate REAL DEFAULT 0,
            line_total REAL DEFAULT 0,
            FOREIGN KEY(invoice_id) REFERENCES purchase_invoices(id),
            FOREIGN KEY(item_id) REFERENCES items(id)
        );
        """
    )

    # Settings
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
    )

    # Add default user if none exists
    if not db.execute("SELECT id FROM users").fetchone():
        db.execute(
            "INSERT INTO users (name, email, password_hash) VALUES (?, ?, ?)",
            ("Admin User", "user@example.com", bcrypt.generate_password_hash("password123").decode("utf-8")),
        )
        db.commit()

# ----------------------
# Auth, session, user
# ----------------------

from functools import wraps

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def load_logged_in_user():
    user_id = session.get('user_id')
    if user_id is None:
        g.user = None
    else:
        db = get_db()
        g.user = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
        if user and bcrypt.check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            flash('Logged in successfully!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# ----------------------
# Core routes
# ----------------------

@app.route('/dashboard')
@login_required
def dashboard():
    db = get_db()

    # Customers
    customer_count = db.execute("SELECT COUNT(*) FROM customers").fetchone()[0]

    # Suppliers
    supplier_count = db.execute("SELECT COUNT(*) FROM suppliers").fetchone()[0]

    # Inventory
    item_count = db.execute("SELECT COUNT(*) FROM items").fetchone()[0]
    stock_value = db.execute("SELECT IFNULL(SUM(cost_price * stock_qty), 0) FROM items").fetchone()[0]
    low_stock = db.execute(
        "SELECT id, name, sku, stock_qty, reorder_level FROM items WHERE stock_qty <= reorder_level"
    ).fetchall()

    # Sales
    sales_count = db.execute("SELECT COUNT(*) FROM sales_invoices").fetchone()[0]
    sales_total = db.execute("SELECT IFNULL(SUM(grand_total), 0) FROM sales_invoices").fetchone()[0]
    outstanding_total = 0

    # Purchases
    purchase_count = db.execute("SELECT COUNT(*) FROM purchase_invoices").fetchone()[0]
    purchase_total = db.execute("SELECT IFNULL(SUM(grand_total), 0) FROM purchase_invoices").fetchone()[0]

    # Stock moves
    stock_moves = db.execute("SELECT COUNT(*) FROM stock_moves").fetchone()[0]

    # Analytics dictionary
    analytics = {
        "counts": {
            "customers": customer_count,
            "suppliers": supplier_count,
            "items": item_count,
            "sales_invoices": sales_count,
            "purchase_invoices": purchase_count,
            "stock_moves": stock_moves,
        },
        "sales_total": sales_total,
        "outstanding_total": outstanding_total,
        "purchase_total": purchase_total,
        "stock_value": stock_value,
        "low_stock": low_stock,
    }

    return render_template(
        "dashboard.html",
        name=session.get("user_name"),
        analytics=analytics,
        low_stock=low_stock
    )

    
    
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')

        if not name or not email or not password or not confirm:
            return render_template('register.html', error="All fields are required.")

        if password != confirm:
            return render_template('register.html', error="Passwords do not match.")

        db = get_db()
        existing_user = db.execute('SELECT id FROM users WHERE email = ?', (email,)).fetchone()
        if existing_user:
            return render_template('register.html', error="Email is already registered.")

        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        db.execute('INSERT INTO users (name, email, password_hash) VALUES (?, ?, ?)',
                   (name, email, password_hash))
        db.commit()

        flash('Registration successful! Please login.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')



# ----------------------
# /api/customers: Get all customers
# ----------------------
@app.route('/api/customers', methods=['GET'])
@login_required
def api_customers():
    db = get_db()
    customers = db.execute('SELECT id, name, mobile, email FROM customers ORDER BY name').fetchall()
    return jsonify(customers)

# ----------------------
# Customers: List
# ----------------------
@app.route('/customers')
@login_required
def customer_list():
    db = get_db()

    # Filters from query params
    q = request.args.get('q', '').strip()
    city = request.args.get('city', '').strip()
    status = request.args.get('status', '').strip()
    per_page = int(request.args.get('per_page', 20))
    page = int(request.args.get('page', 1))

    sql = "SELECT * FROM customers WHERE 1=1"
    params = []

    if q:
        sql += " AND (name LIKE ? OR email LIKE ?)"
        params += [f"%{q}%", f"%{q}%"]
    if city:
        sql += " AND city = ?"
        params.append(city)
    if status:
        sql += " AND status = ?"
        params.append(status)

    total = db.execute(f"SELECT COUNT(*) FROM ({sql})", params).fetchone()[0]
    pages = max(1, -(-total // per_page))  # ceiling division

    sql += " ORDER BY name LIMIT ? OFFSET ?"
    params += [per_page, (page - 1) * per_page]
    customers = db.execute(sql, params).fetchall()

    # For city filter dropdown
    cities = [r[0] for r in db.execute("SELECT DISTINCT city FROM customers WHERE city IS NOT NULL AND city <> ''").fetchall()]

    return render_template(
        'customers/list.html',
        customers=customers,
        q=q,
        city=city,
        status=status,
        per_page=per_page,
        page=page,
        pages=pages,
        total=total,
        cities=cities
    )

# ----------------------
# Customers: New
# ----------------------
@app.route('/customers/new', methods=['GET', 'POST'])
@login_required
def customer_new():
    db = get_db()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        mobile = request.form.get('mobile', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        city = request.form.get('city', '').strip()
        pincode = request.form.get('pincode', '').strip()
        state = request.form.get('state', '').strip()
        country = request.form.get('country', '').strip()
        notes = request.form.get('notes', '').strip()
        status = request.form.get('status', 'active').strip()

        if not name:
            flash("Customer name is required.", "error")
            return redirect(url_for('customers_new'))

        db.execute("""
            INSERT INTO customers (name, email, mobile, phone, address, city, pincode, state, country, notes, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, email, mobile, phone, address, city, pincode, state, country, notes, status))
        db.commit()
        flash("Customer added successfully!", "success")
        return redirect(url_for('customer_list'))

    return render_template('customers/form.html', customer=None)

# ----------------------
# Customers: Edit
# ----------------------
@app.route('/customers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def customer_edit(id):
    db = get_db()
    customer = db.execute("SELECT * FROM customers WHERE id=?", (id,)).fetchone()
    if not customer:
        flash("Customer not found.", "error")
        return redirect(url_for('customers_list'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        mobile = request.form.get('mobile', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        city = request.form.get('city', '').strip()
        pincode = request.form.get('pincode', '').strip()
        state = request.form.get('state', '').strip()
        country = request.form.get('country', '').strip()
        notes = request.form.get('notes', '').strip()
        status = request.form.get('status', 'active').strip()

        db.execute("""
            UPDATE customers
            SET name=?, email=?, mobile=?, phone=?, address=?, city=?, pincode=?, state=?, country=?, notes=?, status=?
            WHERE id=?
        """, (name, email, mobile, phone, address, city, pincode, state, country, notes, status, id))
        db.commit()
        flash("Customer updated successfully!", "success")
        return redirect(url_for('customer_list'))

    return render_template('customers/form.html', customer=customer)

# ----------------------
# Customers: Delete
# ----------------------
@app.route('/customers/<int:id>/delete', methods=['POST'])
@login_required
def customer_delete(id):
    db = get_db()
    db.execute("DELETE FROM customers WHERE id=?", (id,))
    db.commit()
    flash("Customer deleted.", "success")
    return redirect(url_for('customer_list'))



# ----------------------
# Supplier CRUD
# ----------------------

@app.route("/suppliers", endpoint="supplier_list")
def supplier_list():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "")
    per_page = int(request.args.get("per_page", 20))
    page = int(request.args.get("page", 1))

    db = get_db()
    sql = "SELECT * FROM suppliers WHERE 1=1"
    params = []
    if q:
        sql += """ AND (name LIKE ? OR email LIKE ? OR phone LIKE ? OR gstin LIKE ? OR city LIKE ? OR state LIKE ?)"""
        params += [f"%{q}%"] * 6
    if status:
        sql += " AND status = ?"
        params.append(status)

    total = db.execute(f"SELECT COUNT(*) FROM ({sql})", params).fetchone()[0]
    pages = (total + per_page - 1) // per_page
    suppliers = db.execute(
        sql + " ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per_page, (page - 1) * per_page]
    ).fetchall()

    return render_template(
        "suppliers/list.html",
        suppliers=suppliers, q=q, status=status,
        per_page=per_page, page=page, pages=pages, total=total
    )

@app.route("/suppliers/new", methods=["GET", "POST"], endpoint="supplier_new")
def supplier_new():
    if request.method == "POST":
        data = {k: request.form.get(k) for k in [
            "name", "email", "phone", "gstin", "gst_type",
            "address", "city", "pincode", "state",
            "contact_person", "mobile", "status"
        ]}
        db = get_db()
        db.execute("""
            INSERT INTO suppliers 
            (name,email,phone,gstin,gst_type,address,city,pincode,state,contact_person,mobile,status,created_at,updated_at) 
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP,CURRENT_TIMESTAMP)
        """, tuple(data.values()))
        db.commit()
        flash("Supplier created successfully")
        return redirect(url_for("supplier_list"))
    return render_template("suppliers/form.html", supplier=None)

@app.route("/suppliers/<int:id>/edit", methods=["GET", "POST"], endpoint="supplier_edit")
def supplier_edit(id):
    db = get_db()
    supplier = db.execute("SELECT * FROM suppliers WHERE id=?", (id,)).fetchone()
    if not supplier:
        flash("Supplier not found")
        return redirect(url_for("supplier_list"))

    if request.method == "POST":
        data = {k: request.form.get(k) for k in [
            "name", "email", "phone", "gstin", "gst_type",
            "address", "city", "pincode", "state",
            "contact_person", "mobile", "status"
        ]}
        db.execute("""
            UPDATE suppliers SET
              name=?, email=?, phone=?, gstin=?, gst_type=?,
              address=?, city=?, pincode=?, state=?, contact_person=?, mobile=?, status=?,
              updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, tuple(data.values()) + (id,))
        db.commit()
        flash("Supplier updated successfully")
        return redirect(url_for("supplier_list"))

    return render_template("suppliers/form.html", supplier=supplier)

@app.route("/suppliers/<int:id>/delete", methods=["POST"], endpoint="supplier_delete")
def supplier_delete(id):
    db = get_db()
    db.execute("DELETE FROM suppliers WHERE id=?", (id,))
    db.commit()
    flash("Supplier deleted successfully")
    return redirect(url_for("supplier_list"))

# ----------------------
# Import / Export
# ----------------------

@app.route("/suppliers/export.csv")
def supplier_export_csv():
    db = get_db()
    suppliers = db.execute("SELECT * FROM suppliers").fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([c for c in suppliers[0].keys()] if suppliers else [])
    for s in suppliers:
        writer.writerow([s[c] for c in s.keys()])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode("utf-8")),
                     mimetype="text/csv",
                     as_attachment=True,
                     download_name="suppliers.csv")

@app.route("/suppliers/export.xlsx")
def supplier_export_xlsx():
    db = get_db()
    suppliers = db.execute("SELECT * FROM suppliers").fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    if suppliers:
        ws.append([c for c in suppliers[0].keys()])
        for s in suppliers:
            ws.append([s[c] for c in s.keys()])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio,
                     as_attachment=True,
                     download_name="suppliers.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/suppliers/import", methods=["POST"])
def supplier_import():
    file = request.files["file"]
    if not file:
        flash("No file uploaded")
        return redirect(url_for("supplier_list"))

    filename = file.filename
    db = get_db()

    if filename.endswith(".csv"):
        stream = io.StringIO(file.stream.read().decode("utf-8"))
        reader = csv.DictReader(stream)
        for row in reader:
            db.execute("""
                INSERT INTO suppliers (name,email,phone,gstin,gst_type,address,city,pincode,state,contact_person,mobile,status)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                row.get("name"), row.get("email"), row.get("phone"), row.get("gstin"),
                row.get("gst_type","regular"), row.get("address"), row.get("city"),
                row.get("pincode"), row.get("state"), row.get("contact_person"),
                row.get("mobile"), row.get("status","active")
            ))
    elif filename.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rowdict = dict(zip(headers, row))
            db.execute("""
                INSERT INTO suppliers (name,email,phone,gstin,gst_type,address,city,pincode,state,contact_person,mobile,status)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                rowdict.get("name"), rowdict.get("email"), rowdict.get("phone"), rowdict.get("gstin"),
                rowdict.get("gst_type","regular"), rowdict.get("address"), rowdict.get("city"),
                rowdict.get("pincode"), rowdict.get("state"), rowdict.get("contact_person"),
                rowdict.get("mobile"), rowdict.get("status","active")
            ))
    db.commit()
    flash("Suppliers imported successfully")
    return redirect(url_for("supplier_list"))

@app.route("/suppliers/import-template.csv")
def supplier_import_template_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "name","email","phone","gstin","gst_type","address",
        "city","pincode","state","contact_person","mobile","status"
    ])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode("utf-8")),
                     mimetype="text/csv",
                     as_attachment=True,
                     download_name="suppliers_template.csv")

@app.route("/suppliers/import-template.xlsx")
def supplier_import_template_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "name","email","phone","gstin","gst_type","address",
        "city","pincode","state","contact_person","mobile","status"
    ])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio,
                     as_attachment=True,
                     download_name="suppliers_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ----------------------
# /api/items/search: Search for items by SKU or name, used by searchable selects
# ----------------------
@app.route('/api/items/search', methods=['GET'])
@login_required
def api_items_search():
    q = request.args.get('q', '').strip()
    limit = request.args.get('limit', 20, type=int)
    
    db = get_db()
    
    # Build query dynamically
    query = "SELECT id, sku, name, selling_price, stock_qty, uom, discount FROM items WHERE status = 'active'"
    params = []
    
    if q:
        # Search in name, sku, description, brand
        query += " AND (name LIKE ? OR sku LIKE ? OR description LIKE ? OR brand LIKE ?)"
        search_term = f"%{q}%"
        params.extend([search_term, search_term, search_term, search_term])
        
    query += " ORDER BY name LIMIT ?"
    params.append(limit)
    
    items = db.execute(query, params).fetchall()
    
    return jsonify(items)

# ----------------------
# /items: List all inventory items with filters + pagination
# ----------------------
@app.route('/items')
@login_required
def items_list():
    db = get_db()

    # Filters
    q = request.args.get('q', '').strip()
    brand = request.args.get('brand', '').strip()
    supplier = request.args.get('supplier', '').strip()
    status = request.args.get('status', '').strip()

    # Pagination
    per_page = int(request.args.get('per_page', 20))
    page = int(request.args.get('page', 1))

    # Base query
    sql = "SELECT i.*, s.name as supplier_name FROM items i LEFT JOIN suppliers s ON i.supplier_id = s.id WHERE 1=1"
    params = []

    if q:
        sql += " AND (i.name LIKE ? OR i.sku LIKE ? OR i.description LIKE ?)"
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if brand:
        sql += " AND i.brand = ?"
        params.append(brand)
    if supplier:
        sql += " AND s.name = ?"
        params.append(supplier)
    if status:
        sql += " AND i.status = ?"
        params.append(status)

    # Count total
    total = db.execute(sql.replace("i.*, s.name as supplier_name", "COUNT(*)"), params).fetchone()[0]
    pages = max(1, -(-total // per_page))  # ceiling division

    # Fetch paginated results
    sql += " ORDER BY i.name LIMIT ? OFFSET ?"
    params += [per_page, (page - 1) * per_page]
    items = db.execute(sql, params).fetchall()

    # For dropdown filters
    brands = [r['brand'] for r in db.execute("SELECT DISTINCT brand FROM items WHERE brand IS NOT NULL").fetchall()]
    suppliers = [r['name'] for r in db.execute("SELECT DISTINCT name FROM suppliers WHERE name IS NOT NULL").fetchall()]

    return render_template(
        'items/list.html',
        items=items,
        q=q,
        brand=brand,
        supplier=supplier,
        status=status,
        per_page=per_page,
        page=page,
        pages=pages,
        total=total,
        brands=brands,
        suppliers=suppliers
    )


# ----------------------
# /items/new: Create a new item
# ----------------------
@app.route('/items/new', methods=['GET', 'POST'])
@login_required
def item_new():
    db = get_db()
    suppliers = db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()

    if request.method == 'POST':
        form = request.form
        name = form['name'].strip()
        if not name:
            flash("Item name is required", "error")
            return redirect(url_for('item_new'))

        try:
            cursor = db.execute("""
                INSERT INTO items
                (sku, name, description, uom, brand, supplier_id, hsn_code, cost_price, selling_price,
                 gst_rate, discount, stock_qty, reorder_level, status, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (
                form.get('sku') or None,
                name,
                form.get('description'),
                form.get('uom') or 'Nos',
                form.get('brand'),
                form.get('supplier_id') or None,
                form.get('hsn_code'),
                form.get('cost_price') or 0,
                form.get('selling_price') or 0,
                form.get('gst_rate') or 0,
                form.get('discount') or 0,
                form.get('stock_qty') or 0,
                form.get('reorder_level') or 0,
                form.get('status') or 'active'
            ))
            item_id = cursor.lastrowid

            # Auto-generate SKU if blank
            if not form.get('sku'):
                new_sku = f"ITEM{item_id:04d}"
                db.execute("UPDATE items SET sku=? WHERE id=?", (new_sku, item_id))

            db.commit()
            flash("Item created successfully", "success")
            return redirect(url_for('items_list'))
        except Exception as e:
            flash(f"Error creating item: {e}", "error")

    return render_template('items/form.html', item=None, suppliers=suppliers)


# ----------------------
# /items/<id>/edit: Edit existing item
# ----------------------
@app.route('/items/<int:item_id>/edit', methods=['GET', 'POST'])
@login_required
def item_edit(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()
    suppliers = db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()

    if not item:
        flash("Item not found", "error")
        return redirect(url_for('items_list'))

    if request.method == 'POST':
        form = request.form
        try:
            db.execute("""
                UPDATE items SET
                sku=?, name=?, description=?, uom=?, brand=?, supplier_id=?, hsn_code=?,
                cost_price=?, selling_price=?, gst_rate=?, discount=?, stock_qty=?, reorder_level=?,
                status=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, (
                form.get('sku') or None,
                form['name'].strip(),
                form.get('description'),
                form.get('uom') or 'Nos',
                form.get('brand'),
                form.get('supplier_id') or None,
                form.get('hsn_code'),
                form.get('cost_price') or 0,
                form.get('selling_price') or 0,
                form.get('gst_rate') or 0,
                form.get('discount') or 0,
                form.get('stock_qty') or 0,
                form.get('reorder_level') or 0,
                form.get('status') or 'active',
                item_id
            ))
            db.commit()
            flash("Item updated successfully", "success")
            return redirect(url_for('items_list'))
        except Exception as e:
            flash(f"Error updating item: {e}", "error")

    return render_template('items/form.html', item=item, suppliers=suppliers)


# ----------------------
# /items/<id>/delete: Delete an item
# ----------------------
@app.route('/items/<int:item_id>/delete', methods=['POST'])
@login_required
def item_delete(item_id):
    db = get_db()
    try:
        db.execute("DELETE FROM items WHERE id=?", (item_id,))
        db.commit()
        flash("Item deleted successfully", "success")
    except Exception as e:
        flash(f"Error deleting item: {e}", "error")
    return redirect(url_for('items_list'))

import io
import csv
import openpyxl
from flask import send_file, request, redirect, url_for, flash, Response

# ----------------------
# Export Items (CSV)
# ----------------------
@app.route('/items/export.csv')
@login_required
def items_export_csv():
    db = get_db()
    items = db.execute("""
        SELECT sku, name, description, uom, brand, hsn_code, cost_price,
               selling_price, gst_rate, discount, stock_qty, reorder_level, status
        FROM items ORDER BY name
    """).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    headers = ["SKU", "Name", "Description", "UOM", "Brand", "HSN Code",
               "Cost Price", "Selling Price", "GST Rate", "Discount",
               "Stock Qty", "Reorder Level", "Status"]
    writer.writerow(headers)

    for i in items:
        writer.writerow([
            i["sku"], i["name"], i["description"], i["uom"], i["brand"], i["hsn_code"],
            i["cost_price"], i["selling_price"], i["gst_rate"], i["discount"],
            i["stock_qty"], i["reorder_level"], i["status"]
        ])

    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=items.csv"
    return response


# ----------------------
# Export Items (XLSX)
# ----------------------
@app.route('/items/export.xlsx')
@login_required
def items_export_xlsx():
    db = get_db()
    items = db.execute("""
        SELECT sku, name, description, uom, brand, hsn_code, cost_price,
               selling_price, gst_rate, discount, stock_qty, reorder_level, status
        FROM items ORDER BY name
    """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    headers = ["SKU", "Name", "Description", "UOM", "Brand", "HSN Code",
               "Cost Price", "Selling Price", "GST Rate", "Discount",
               "Stock Qty", "Reorder Level", "Status"]
    ws.append(headers)

    for i in items:
        ws.append([
            i["sku"], i["name"], i["description"], i["uom"], i["brand"], i["hsn_code"],
            i["cost_price"], i["selling_price"], i["gst_rate"], i["discount"],
            i["stock_qty"], i["reorder_level"], i["status"]
        ])

    file_io = io.BytesIO()
    wb.save(file_io)
    file_io.seek(0)

    return send_file(
        file_io,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="items.xlsx"
    )


# ----------------------
# Import Templates (CSV/XLSX)
# ----------------------
@app.route('/items/import-template.csv')
@login_required
def items_import_template_csv():
    headers = ["SKU", "Name", "Description", "UOM", "Brand", "HSN Code",
               "Cost Price", "Selling Price", "GST Rate", "Discount",
               "Stock Qty", "Reorder Level", "Status"]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=items_import_template.csv"
    return response


@app.route('/items/import-template.xlsx')
@login_required
def items_import_template_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    headers = ["SKU", "Name", "Description", "UOM", "Brand", "HSN Code",
               "Cost Price", "Selling Price", "GST Rate", "Discount",
               "Stock Qty", "Reorder Level", "Status"]
    ws.append(headers)

    file_io = io.BytesIO()
    wb.save(file_io)
    file_io.seek(0)

    return send_file(
        file_io,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="items_import_template.xlsx"
    )


# ----------------------
# Import Items (CSV/XLSX) with duplicate handling
# ----------------------
@app.route('/items/import', methods=['POST'])
@login_required
def items_import():
    file = request.files.get('file')
    if not file:
        flash("No file uploaded", "error")
        return redirect(url_for('items_list'))

    filename = file.filename.lower()
    db = get_db()
    rows = []

    try:
        if filename.endswith(".csv"):
            stream = io.StringIO(file.stream.read().decode("utf-8"))
            reader = csv.DictReader(stream)
            rows = list(reader)

        elif filename.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # skip empty rows
                    continue
                rows.append(dict(zip(headers, row)))

        else:
            flash("Unsupported file type", "error")
            return redirect(url_for('items_list'))

        count = 0
        for row in rows:
            if not row.get("Name"):
                continue  # skip incomplete rows

            db.execute("""
                INSERT INTO items
                (sku, name, description, uom, brand, hsn_code, cost_price,
                 selling_price, gst_rate, discount, stock_qty, reorder_level, status, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(sku) DO UPDATE SET
                  name=excluded.name,
                  description=excluded.description,
                  uom=excluded.uom,
                  brand=excluded.brand,
                  hsn_code=excluded.hsn_code,
                  cost_price=excluded.cost_price,
                  selling_price=excluded.selling_price,
                  gst_rate=excluded.gst_rate,
                  discount=excluded.discount,
                  stock_qty=excluded.stock_qty,
                  reorder_level=excluded.reorder_level,
                  status=excluded.status,
                  updated_at=CURRENT_TIMESTAMP
            """, (
                row.get("SKU"), row.get("Name"), row.get("Description"),
                row.get("UOM") or "Nos", row.get("Brand"), row.get("HSN Code"),
                row.get("Cost Price") or 0, row.get("Selling Price") or 0,
                row.get("GST Rate") or 0, row.get("Discount") or 0,
                row.get("Stock Qty") or 0, row.get("Reorder Level") or 0,
                row.get("Status") or "active"
            ))
            count += 1

        db.commit()
        flash(f"Imported {count} items successfully", "success")

    except Exception as e:
        flash(f"Import failed: {e}", "error")

    return redirect(url_for('items_list'))

# ----------------------
# Items-- Barcode 
# ----------------------
import io, math, zipfile
from flask import send_file, render_template, request, redirect, url_for, flash
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont

import io, math, zipfile
from flask import send_file, render_template, request, redirect, url_for, flash
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont

@app.route('/items/barcodes', methods=['GET', 'POST'])
@login_required
def items_barcodes():
    db = get_db()

    if request.method == 'POST':
        item_ids = request.form.getlist('item_id')
        if not item_ids:
            flash("No items selected", "error")
            return redirect(url_for('items_barcodes'))

        # Collect barcode entries
        barcodes = []
        for item_id in item_ids:
            copies = int(request.form.get(f'qty[{item_id}]', 1))
            item = db.execute(
                "SELECT sku, name, selling_price, stock_qty, uom FROM items WHERE id=?",
                (item_id,)
            ).fetchone()
            if not item:
                continue

            # Generate barcode
            code128 = Code128(item['sku'], writer=ImageWriter())
            img_io = io.BytesIO()
            code128.write(img_io, {"module_height": 25.0, "font_size": 10})
            img_io.seek(0)
            barcode_img = Image.open(img_io).convert("RGB")

            for _ in range(copies):
                barcodes.append({
                    "img": barcode_img,
                    "name": item["name"],
                    "sku": item["sku"],
                    "price": item["selling_price"],
                    "stock_qty": item["stock_qty"],
                    "uom": item["uom"] or ""   # fallback empty if null
                })

        if not barcodes:
            flash("No barcodes generated", "error")
            return redirect(url_for('items_barcodes'))

        # Layout constants
        DPI = 300
        PAGE_W, PAGE_H = 2480, 3508  # A4 @ 300dpi
        COLS = 5
        MARGIN_X, MARGIN_Y = 50, 50
        GAP_X, GAP_Y = 30, 40
        CELL_W = (PAGE_W - 2*MARGIN_X - (COLS-1)*GAP_X) // COLS
        CELL_H = 320  # barcode + text

        ROWS = (PAGE_H - 2*MARGIN_Y) // (CELL_H + GAP_Y)
        per_page = COLS * ROWS
        total_pages = math.ceil(len(barcodes) / per_page)

        # Fonts
        try:
            font_title = ImageFont.truetype("DejaVuSans-Bold.ttf", 24)   # header
            font_text = ImageFont.truetype("DejaVuSans.ttf", 18)        # normal text
            font_price = ImageFont.truetype("DejaVuSans-Bold.ttf", 26)  # bigger for price
        except:
            font_title = font_text = font_price = ImageFont.load_default()

        # ZIP output
        mem_zip = io.BytesIO()
        with zipfile.ZipFile(mem_zip, "w") as zf:
            for page_idx in range(total_pages):
                page = Image.new("RGB", (PAGE_W, PAGE_H), "white")
                draw = ImageDraw.Draw(page)

                start = page_idx * per_page
                end = min(start + per_page, len(barcodes))

                for i, bc in enumerate(barcodes[start:end]):
                    row = i // COLS
                    col = i % COLS
                    x = MARGIN_X + col * (CELL_W + GAP_X)
                    y = MARGIN_Y + row * (CELL_H + GAP_Y)

                    # Brand header - centered
                    header = "FLORA TEXTILE BI"
                    w, h = draw.textsize(header, font=font_title)
                    text_x = x + (CELL_W - w) // 2
                    draw.text((text_x, y), header, font=font_title, fill="black")

                    # Barcode (below header)
                    bc_resized = bc["img"].resize((CELL_W, 100), Image.LANCZOS)
                    page.paste(bc_resized, (x, y + 30))

                    # Item name - shrink font if needed, center
                    name = bc["name"] or ""
                    font_name = font_text
                    max_width = CELL_W - 10
                    w, h = draw.textsize(name, font=font_name)
                    while w > max_width and font_name.size > 12:
                        font_name = ImageFont.truetype("DejaVuSans.ttf", font_name.size - 2)
                        w, h = draw.textsize(name, font=font_name)
                    text_x = x + (CELL_W - w) // 2
                    draw.text((text_x, y + 140), name, font=font_name, fill="black")

                    # SKU - centered
                    sku_text = f"SKU: {bc['sku']}"
                    w, h = draw.textsize(sku_text, font=font_text)
                    text_x = x + (CELL_W - w) // 2
                    draw.text((text_x, y + 170), sku_text, font=font_text, fill="black")

                    # Selling Price - bold & bigger
                    price_text = f"₹ {bc['price']:.2f}"
                    w, h = draw.textsize(price_text, font=font_price)
                    text_x = x + (CELL_W - w) // 2
                    draw.text((text_x, y + 200), price_text, font=font_price, fill="black")

                    # Stock Qty - centered
                    qty_text = f"Stock: {bc['stock_qty']} {bc['uom']}"
                    w, h = draw.textsize(qty_text, font=font_text)
                    text_x = x + (CELL_W - w) // 2
                    draw.text((text_x, y + 240), qty_text, font=font_text, fill="black")

                # Save PNG
                img_io = io.BytesIO()
                page.save(img_io, format="PNG", dpi=(DPI, DPI))
                zf.writestr(f"barcodes_page_{page_idx+1}.png", img_io.getvalue())

        mem_zip.seek(0)
        return send_file(
            mem_zip,
            mimetype="application/zip",
            as_attachment=True,
            download_name="barcodes.zip"
        )

    # GET → selection form
    items = db.execute("""
        SELECT items.*, suppliers.name as supplier_name
        FROM items LEFT JOIN suppliers ON suppliers.id = items.supplier_id
        ORDER BY items.name
    """).fetchall()
    return render_template('items/barcodes.html', items=items)



# ----------------------
# Sales Orders CRUD (complete, robust)
# ----------------------


# ---------- helpers ----------
def safe_float(val):
    """Convert string (with optional comma) to float safely."""
    try:
        if val is None or str(val).strip() == "":
            return 0.0
        return float(str(val).replace(',', '.'))
    except Exception:
        return 0.0

def calculate_line(qty, rate, disc):
    """Return normalized qty, rate, disc and computed line_total."""
    qty = safe_float(qty)
    rate = safe_float(rate)
    disc = safe_float(disc)
    base = qty * rate
    line_total = max(base * (1 - disc / 100), 0)
    return qty, rate, disc, line_total

# --- Sales Order List ---
@app.route("/sales-orders")
@login_required
def salesorder_list():
    db = get_db()

    # Filters & search
    search = request.args.get("search", "").strip()
    customer_id = request.args.get("customer_id", "")
    page = int(request.args.get("page", 1))
    per_page = 10
    offset = (page - 1) * per_page

    query = """
        SELECT so.id, so.so_no, so.date, so.expected_delivery_date, 
               so.total, so.grand_total, c.name AS customer_name
        FROM sales_orders so
        LEFT JOIN customers c ON so.customer_id = c.id
        WHERE 1=1
    """
    params = []

    if search:
        query += " AND (so.so_no LIKE ? OR c.name LIKE ?)"
        params.extend([f"%{search}%", f"%{search}%"])

    if customer_id:
        query += " AND so.customer_id = ?"
        params.append(customer_id)

    query += " ORDER BY so.date DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])

    sales_orders = db.execute(query, params).fetchall()

    # Count for pagination
    count_query = """
        SELECT COUNT(*) 
        FROM sales_orders so
        LEFT JOIN customers c ON so.customer_id = c.id
        WHERE 1=1
    """
    count_params = []

    if search:
        count_query += " AND (so.so_no LIKE ? OR c.name LIKE ?)"
        count_params.extend([f"%{search}%", f"%{search}%"])

    if customer_id:
        count_query += " AND so.customer_id = ?"
        count_params.append(customer_id)

    total_count = db.execute(count_query, count_params).fetchone()[0]
    total_pages = (total_count + per_page - 1) // per_page

    customers = db.execute("SELECT id, name FROM customers ORDER BY name").fetchall()

    return render_template(
        "sales_orders/list.html",
        sales_orders=sales_orders,
        page=page,
        total_pages=total_pages,
        search=search,
        customer_id=customer_id,
        customers=customers
    )



# --- New Sales Order ---
@app.route("/sales-orders/new", methods=["GET", "POST"], endpoint="salesorder_new")
@login_required
def salesorder_new():
    db = get_db()
    if request.method == "POST":
        customer_id = request.form.get("customer_id")
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        expected_delivery_date = request.form.get("expected_delivery_date")
        notes = request.form.get("notes")
        qtys = request.form.getlist("qty[]")
        rates = request.form.getlist("rate[]")
        discounts = request.form.getlist("discount[]")
        item_ids = request.form.getlist("item_id[]")

        total = 0.0

        # 1. Insert SO placeholder (so_no blank for now)
        cur = db.execute("""
            INSERT INTO sales_orders (customer_id, so_no, date, expected_delivery_date, notes, total, grand_total)
            VALUES (?, '', ?, ?, ?, 0, 0)
        """, (customer_id, date, expected_delivery_date, notes))
        so_id = cur.lastrowid

         # 2. Generate SO number with daily reset
        today_str = datetime.now().strftime("%Y%m%d")

        # Count how many SOs already exist for today
        count_today = db.execute("""
            SELECT COUNT(*) 
            FROM sales_orders 
            WHERE so_no LIKE ?
        """, (f"SO{today_str}-%",)).fetchone()[0]

        next_seq = count_today + 1
        so_no = f"SO{today_str}-{next_seq:05d}"

        # 3. Update order with generated SO number
        db.execute("UPDATE sales_orders SET so_no=? WHERE id=?", (so_no, so_id))

        
        # 4. Insert items
        for i in range(len(item_ids)):
            if not item_ids[i]:
                continue
            qty, rate, disc, line_total = calculate_line(qtys[i], rates[i], discounts[i])
            total += line_total
            db.execute("""
                INSERT INTO sales_order_items (sales_order_id, item_id, qty, rate, discount, line_total)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (so_id, item_ids[i], qty, rate, disc, line_total))

        # 5. Update totals
        db.execute("UPDATE sales_orders SET total=?, grand_total=? WHERE id=?", (total, total, so_id))
        db.commit()

        flash(f"Sales Order {so_no} created successfully!", "success")
        return redirect(url_for("salesorder_list"))

    # GET → render form
    customers = db.execute("SELECT id, name FROM customers ORDER BY name").fetchall()
    return render_template(
        "sales_orders/form.html",
        so=None,
        lines=[],
        customers=customers,
        selected_customer=None,
        date_today=datetime.today().strftime("%Y-%m-%d"),
    )

# --- Edit Sales Order ---
@app.route("/sales-orders/<int:so_id>/edit", methods=["GET", "POST"])
@login_required
def salesorder_edit(so_id):
    db = get_db()
    so = db.execute("SELECT * FROM sales_orders WHERE id=?", (so_id,)).fetchone()
    if not so:
        flash("Sales Order not found", "danger")
        return redirect(url_for("salesorder_list"))

    if request.method == "POST":
        customer_id = request.form.get("customer_id")
        so_no = request.form.get("so_no")
        date = request.form.get("date")
        expected_delivery_date = request.form.get("expected_delivery_date")
        notes = request.form.get("notes")
        qtys = request.form.getlist("qty[]")
        rates = request.form.getlist("rate[]")
        discounts = request.form.getlist("discount[]")
        item_ids = request.form.getlist("item_id[]")

        db.execute("DELETE FROM sales_order_items WHERE sales_order_id=?", (so_id,))

        total = 0.0
        for i in range(len(item_ids)):
            if not item_ids[i]:
                continue
            qty, rate, disc, line_total = calculate_line(qtys[i], rates[i], discounts[i])
            total += line_total
            db.execute("""
                INSERT INTO sales_order_items (sales_order_id, item_id, qty, rate, discount, line_total)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (so_id, item_ids[i], qty, rate, disc, line_total))

        db.execute("""
            UPDATE sales_orders
            SET customer_id=?, so_no=?, date=?, expected_delivery_date=?, notes=?, total=?, grand_total=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (customer_id, so_no, date, expected_delivery_date, notes, total, total, so_id))
        db.commit()
        flash("Sales Order updated successfully!", "success")
        return redirect(url_for("salesorder_list"))

    lines = db.execute("""
        SELECT l.id, l.item_id, l.qty, l.rate, l.discount, l.line_total,
               i.name as item_name, i.sku as item_sku
        FROM sales_order_items l
        JOIN items i ON l.item_id=i.id
        WHERE l.sales_order_id=?
    """, (so_id,)).fetchall()
    customers = db.execute("SELECT id, name FROM customers ORDER BY name").fetchall()

    return render_template("sales_orders/form.html",
                           so=so,
                           lines=[dict(l) for l in lines],
                           customers=customers,
                           selected_customer=so["customer_id"],
                           date_today=datetime.today().strftime("%Y-%m-%d"))


# --- Delete Sales Order ---
@app.route("/sales-orders/<int:id>/delete", methods=["POST"])
@login_required
def salesorder_delete(id):
    db = get_db()
    db.execute("DELETE FROM sales_order_items WHERE sales_order_id=?", (id,))
    db.execute("DELETE FROM sales_orders WHERE id=?", (id,))
    db.commit()
    flash("Sales Order deleted.", "success")
    return redirect(url_for("salesorder_list"))


# --- Helper function for pdf and print ---
def get_salesorder_with_lines(id):
    db = get_db()

    # Fetch SO master
    so = db.execute("""
        SELECT so.*, c.name as customer_name, c.address as customer_address
        FROM sales_orders so
        LEFT JOIN customers c ON so.customer_id = c.id
        WHERE so.id = ?
    """, (id,)).fetchone()

    if not so:
        return None, None

    # Fetch SO line items with SKU, Name, HSN
    lines = db.execute("""
        SELECT l.id, l.item_id, l.qty, l.rate, l.discount, l.line_total,
               i.sku AS item_sku,
               i.name AS item_name,
               i.hsn_code AS hsn_code
        FROM sales_order_items l
        JOIN items i ON l.item_id = i.id
        WHERE l.sales_order_id = ?
    """, (id,)).fetchall()

    return so, lines


# --- Print Sales Order (HTML preview) ---
@app.route("/sales-orders/<int:id>/print", methods=["GET"], endpoint="salesorder_print")
@login_required
def salesorder_print(id):
    so, lines = get_salesorder_with_lines(id)

    if not so:
        flash("Sales Order not found.", "danger")
        return redirect(url_for("salesorder_list"))

    return render_template("print/sales_order.html", so=so, lines=lines)


# --- PDF Export ---
@app.route("/sales-orders/<int:id>/pdf", methods=["GET"], endpoint="salesorder_pdf")
@login_required
def salesorder_pdf(id):
    so, lines = get_salesorder_with_lines(id)

    if not so:
        flash("Sales Order not found.", "danger")
        return redirect(url_for("salesorder_list"))

    html = render_template("print/sales_order.html", so=so, lines=lines)

    pdf_file = BytesIO()
    HTML(string=html, base_url=request.root_path).write_pdf(pdf_file)
    pdf_file.seek(0)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=f"SalesOrder_{id}.pdf",
        mimetype="application/pdf"
    )




# ----------------------
# Sales Invoices CRUD + Convert from SO to SI (with Pagination)
# ----------------------
@app.route('/sales-invoices')
@login_required
def salesinvoice_list():
    db = get_db()
    page = int(request.args.get("page", 1))
    per_page = 20
    offset = (page - 1) * per_page

    total = db.execute("SELECT COUNT(*) as count FROM sales_invoices").fetchone()["count"]

    sql = """
        SELECT si.*, c.name as customer_name, so.so_no
        FROM sales_invoices si
        JOIN customers c ON si.customer_id = c.id
        LEFT JOIN sales_orders so ON si.sales_order_id = so.id
        ORDER BY si.date DESC, si.id DESC
        LIMIT ? OFFSET ?
    """
    invoices = db.execute(sql, (per_page, offset)).fetchall()
    total_pages = (total + per_page - 1) // per_page

    return render_template("sales_invoices/list.html",
                           invoices=invoices,
                           page=page,
                           total_pages=total_pages)

@app.route('/sales-invoices/new', methods=['GET', 'POST'])
@login_required
def salesinvoice_new():
    db = get_db()
    customers = db.execute("SELECT id, name FROM customers ORDER BY name").fetchall()
    items = db.execute("SELECT id, name, sku, selling_price, discount FROM items WHERE status='active' ORDER BY name").fetchall()
    if request.method == 'POST':
        return save_invoice(db, request.form)
    return render_template('sales_invoices/form.html', invoice=None, lines=[], customers=customers, items=items,
                           date_today=datetime.today().strftime("%Y-%m-%d"))

@app.route('/sales-invoices/from-so/<int:so_id>')
@login_required
def salesinvoice_from_so(so_id):
    db = get_db()
    so = db.execute("SELECT * FROM sales_orders WHERE id=?", (so_id,)).fetchone()
    lines = db.execute("SELECT * FROM sales_order_items WHERE sales_order_id=?", (so_id,)).fetchall()
    customers = db.execute("SELECT id, name FROM customers ORDER BY name").fetchall()
    items = db.execute("SELECT id, name, sku, selling_price, discount FROM items WHERE status='active' ORDER BY name").fetchall()
    return render_template("sales_invoices/form.html",
                           invoice=None,
                           lines=lines,
                           customers=customers,
                           items=items,
                           so=so,
                           date_today=datetime.today().strftime("%Y-%m-%d"))

def save_invoice(db, form, id=None):
    customer_id = form.get('customer_id')
    date = form.get('date')
    due_date = form.get('due_date')
    invoice_no = form.get('invoice_no') or f"INV{int(datetime.now().timestamp())}"
    notes = form.get('notes')
    so_id = form.get('sales_order_id')

    item_ids = form.getlist('item_id[]')
    qtys = form.getlist('qty[]')
    rates = form.getlist('rate[]')
    discounts = form.getlist('discount[]')

    total = 0
    line_items = []
    for i in range(len(item_ids)):
        qty = float(qtys[i] or 0)
        rate = float(rates[i] or 0)
        disc = float(discounts[i] or 0)
        base = qty * rate
        line_total = max(base * (1 - disc/100), 0)
        total += line_total
        line_items.append((item_ids[i], qty, rate, disc, line_total))

    if id is None:
        cursor = db.execute("""
            INSERT INTO sales_invoices (invoice_no, sales_order_id, customer_id, date, due_date, notes, total, grand_total)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (invoice_no, so_id, customer_id, date, due_date, notes, total, total))
        invoice_id = cursor.lastrowid
    else:
        db.execute("""
            UPDATE sales_invoices
            SET sales_order_id=?, customer_id=?, date=?, due_date=?, invoice_no=?, notes=?, total=?, grand_total=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (so_id, customer_id, date, due_date, invoice_no, notes, total, total, id))
        db.execute("DELETE FROM sales_invoice_items WHERE sales_invoice_id=?", (id,))
        invoice_id = id

    for item_id, qty, rate, disc, line_total in line_items:
        db.execute("""
            INSERT INTO sales_invoice_items (sales_invoice_id, item_id, qty, rate, discount, line_total)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (invoice_id, item_id, qty, rate, disc, line_total))

    db.commit()
    flash("Sales Invoice saved successfully!", "success")
    return redirect(url_for('salesinvoice_list'))

@app.route('/sales-invoices/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def salesinvoice_edit(id):
    db = get_db()
    invoice = db.execute("SELECT * FROM sales_invoices WHERE id=?", (id,)).fetchone()
    lines = db.execute("SELECT * FROM sales_invoice_items WHERE sales_invoice_id=?", (id,)).fetchall()
    customers = db.execute("SELECT id, name FROM customers ORDER BY name").fetchall()
    items = db.execute("SELECT id, name, sku, selling_price, discount FROM items WHERE status='active' ORDER BY name").fetchall()
    if request.method == 'POST':
        return save_invoice(db, request.form, id=id)
    return render_template('sales_invoices/form.html', invoice=invoice, lines=lines, customers=customers, items=items,
                           date_today=datetime.today().strftime("%Y-%m-%d"))

@app.route('/sales-invoices/<int:id>/delete', methods=['POST'])
@login_required
def salesinvoice_delete(id):
    db = get_db()
    db.execute("DELETE FROM sales_invoice_items WHERE sales_invoice_id=?", (id,))
    db.execute("DELETE FROM sales_invoices WHERE id=?", (id,))
    db.commit()
    flash("Invoice deleted.", "success")
    return redirect(url_for('salesinvoice_list'))


# --- API Routes for Items (autocomplete / search)
@app.route("/api/items/search")
@login_required
def search_items():
    q = request.args.get("q", "").strip()
    limit = int(request.args.get("limit", 10))
    page = int(request.args.get("page", 1))
    offset = (page - 1) * limit
    context = request.args.get("context", "")

    db = get_db()
    rows = db.execute(
        """
        SELECT
            id,
            name,
            sku,
            IFNULL(selling_price, 0) AS selling_price,
            IFNULL(discount, 0) AS discount,
            IFNULL(stock_qty, 0) AS stock_qty,
            IFNULL(uom, 'Nos') AS uom
        FROM items
        WHERE name LIKE ? COLLATE NOCASE
           OR sku LIKE ? COLLATE NOCASE
        ORDER BY name
        LIMIT ? OFFSET ?
        """,
        (f"%{q}%", f"%{q}%", limit, offset)
    ).fetchall()

    return jsonify({"results": [dict(r) for r in rows]})




# --- API Routes for customers (autocomplete / search)
@app.route("/api/customers/search")
@login_required
def search_customers():
    q = request.args.get("q", "").strip()
    limit = int(request.args.get("limit", 10))
    page = int(request.args.get("page", 1))
    offset = (page - 1) * limit

    db = get_db()
    rows = db.execute(
        """SELECT id, name, email, phone
           FROM customers
           WHERE name LIKE ? OR email LIKE ? OR phone LIKE ?
           ORDER BY name
           LIMIT ? OFFSET ?""",
        (f"%{q}%", f"%{q}%", f"%{q}%", limit, offset)
    ).fetchall()

    return jsonify({"results": [dict(r) for r in rows]})


# ----------------------
# Purchase Orders CRUD
# ----------------------

# ---------- helpers ----------
def safe_float(val):
    try:
        if val is None or str(val).strip() == "":
            return 0.0
        return float(str(val).replace(',', '.'))
    except Exception:
        return 0.0


def calculate_line(qty, rate, disc, gst):
    """Return normalized values and computed line_total & tax."""
    qty = safe_float(qty)
    rate = safe_float(rate)
    disc = safe_float(disc)
    gst = safe_float(gst)

    base = qty * rate
    net = max(base - disc, 0)
    tax = net * (gst / 100)
    line_total = net + tax
    return qty, rate, disc, gst, net, tax, line_total




@app.route("/purchase-orders", endpoint="po_list")
@login_required
def po_list():
    db = get_db()
    search = request.args.get("search", "").strip()
    supplier_id = request.args.get("supplier_id", "")
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 10))
    offset = (page - 1) * per_page

    query = """
        SELECT po.id, po.po_no, po.date, po.status, po.grand_total,
               s.name AS supplier_name
        FROM purchase_orders po
        LEFT JOIN suppliers s ON po.supplier_id = s.id
        WHERE 1=1
    """
    params = []

    if search:
        query += " AND (po.po_no LIKE ? OR s.name LIKE ?)"
        params.extend([f"%{search}%", f"%{search}%"])

    if supplier_id:
        query += " AND po.supplier_id = ?"
        params.append(supplier_id)

    query += " ORDER BY po.date DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])
    purchase_orders = db.execute(query, params).fetchall()

    # Count for pagination
    count_query = "SELECT COUNT(*) FROM purchase_orders po LEFT JOIN suppliers s ON po.supplier_id = s.id WHERE 1=1"
    count_params = []
    if search:
        count_query += " AND (po.po_no LIKE ? OR s.name LIKE ?)"
        count_params.extend([f"%{search}%", f"%{search}%"])
    if supplier_id:
        count_query += " AND po.supplier_id = ?"
        count_params.append(supplier_id)

    total_count = db.execute(count_query, count_params).fetchone()[0]
    total_pages = ceil(total_count / per_page)

    # Build simple pagination object
    class Pagination:
        def __init__(self, page, per_page, total):
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = ceil(total / per_page)

        @property
        def has_prev(self):
            return self.page > 1

        @property
        def has_next(self):
            return self.page < self.pages

        @property
        def prev_num(self):
            return self.page - 1

        @property
        def next_num(self):
            return self.page + 1

        def iter_pages(self, left_edge=2, left_current=2, right_current=2, right_edge=2):
            last = 0
            for num in range(1, self.pages + 1):
                if num <= left_edge or \
                   (self.page - left_current <= num <= self.page + right_current) or \
                   num > self.pages - right_edge:
                    if last + 1 != num:
                        yield None
                    yield num
                    last = num

    pagination = Pagination(page, per_page, total_count)

    suppliers = db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()

    return render_template(
        "purchase_orders/list.html",
        purchase_orders=purchase_orders,
        suppliers=suppliers,
        search=search,
        supplier_id=supplier_id,
        per_page=per_page,
        pagination=pagination
    )


# --- New ---
@app.route("/purchase-orders/new", methods=["GET", "POST"], endpoint="po_new")
@login_required
def po_new():
    db = get_db()
    if request.method == "POST":
        supplier_id = request.form.get("supplier_id")
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        po_no = request.form.get("po_no")
        notes = request.form.get("notes")

        qtys = request.form.getlist("qty[]")
        rates = request.form.getlist("rate[]")
        discounts = request.form.getlist("discount[]")
        gst_rates = request.form.getlist("gst_rate[]")
        item_ids = request.form.getlist("item_id[]")

        total, tax_total, grand_total = 0.0, 0.0, 0.0

        # Insert placeholder PO
        cur = db.execute("""
            INSERT INTO purchase_orders (supplier_id, po_no, date, notes, total, tax_total, grand_total)
            VALUES (?, '', ?, ?, 0, 0, 0)
        """, (supplier_id, date, notes))
        po_id = cur.lastrowid

        # Auto-generate PO number if empty
        if not po_no:
            today_str = datetime.now().strftime("%Y%m%d")
            count_today = db.execute(
                "SELECT COUNT(*) FROM purchase_orders WHERE po_no LIKE ?",
                (f"PO{today_str}-%",)
            ).fetchone()[0]
            po_no = f"PO{today_str}-{count_today + 1:05d}"

        db.execute("UPDATE purchase_orders SET po_no=? WHERE id=?", (po_no, po_id))

        # Insert line items
        for i in range(len(item_ids)):
            if not item_ids[i]:
                continue
            qty, rate, disc, gst, net, tax, line_total = calculate_line(
                qtys[i], rates[i], discounts[i], gst_rates[i]
            )
            total += net
            tax_total += tax
            grand_total += line_total
            db.execute("""
                INSERT INTO purchase_order_items 
                (purchase_order_id, item_id, qty, rate, discount, gst_rate, net, tax, line_total)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (po_id, item_ids[i], qty, rate, disc, gst, net, tax, line_total))

        db.execute(
            "UPDATE purchase_orders SET total=?, tax_total=?, grand_total=? WHERE id=?",
            (total, tax_total, grand_total, po_id)
        )
        db.commit()
        flash(f"Purchase Order {po_no} created successfully!", "success")
        return redirect(url_for("po_list"))

    suppliers = db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()
    
    return render_template(
        "purchase_orders/form.html",
        po=None,
        lines=[],
        suppliers=suppliers,
        date_today=datetime.today().strftime("%Y-%m-%d"),
    )


# --- Edit ---
@app.route("/purchase-orders/<int:po_id>/edit", methods=["GET", "POST"], endpoint="po_edit")
@login_required
def po_edit(po_id):
    db = get_db()
    po = db.execute("SELECT * FROM purchase_orders WHERE id=?", (po_id,)).fetchone()
    if not po:
        flash("Purchase Order not found", "danger")
        return redirect(url_for("po_list"))

    if request.method == "POST":
        supplier_id = request.form.get("supplier_id")
        po_no = request.form.get("po_no")
        date = request.form.get("date")
        notes = request.form.get("notes")

        qtys = request.form.getlist("qty[]")
        rates = request.form.getlist("rate[]")
        discounts = request.form.getlist("discount[]")
        gst_rates = request.form.getlist("gst_rate[]")
        item_ids = request.form.getlist("item_id[]")

        db.execute("DELETE FROM purchase_order_items WHERE purchase_order_id=?", (po_id,))

        total, tax_total, grand_total = 0.0, 0.0, 0.0
        for i in range(len(item_ids)):
            if not item_ids[i]:
                continue
            qty, rate, disc, gst, net, tax, line_total = calculate_line(
                qtys[i], rates[i], discounts[i], gst_rates[i]
            )
            total += net
            tax_total += tax
            grand_total += line_total
            db.execute("""
                INSERT INTO purchase_order_items 
                (purchase_order_id, item_id, qty, rate, discount, gst_rate, net, tax, line_total)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (po_id, item_ids[i], qty, rate, disc, gst, net, tax, line_total))

        db.execute("""
            UPDATE purchase_orders
            SET supplier_id=?, po_no=?, date=?, notes=?, 
                total=?, tax_total=?, grand_total=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (supplier_id, po_no, date, notes, total, tax_total, grand_total, po_id))
        db.commit()
        flash("Purchase Order updated successfully!", "success")
        return redirect(url_for("po_list"))

    lines = db.execute("""
        SELECT l.*, i.name as item_name, i.sku as item_sku
        FROM purchase_order_items l
        JOIN items i ON l.item_id=i.id
        WHERE l.purchase_order_id=?
    """, (po_id,)).fetchall()
    suppliers = db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()

    return render_template(
        "purchase_orders/form.html",
        po=po,
        lines=[dict(l) for l in lines],
        suppliers=suppliers,
        date_today=datetime.today().strftime("%Y-%m-%d"),
    )


# --- Delete ---
@app.route("/purchase-orders/<int:id>/delete", methods=["POST"], endpoint="po_delete")
@login_required
def po_delete(id):
    db = get_db()
    db.execute("DELETE FROM purchase_order_items WHERE purchase_order_id=?", (id,))
    db.execute("DELETE FROM purchase_orders WHERE id=?", (id,))
    db.commit()
    flash("Purchase Order deleted.", "success")
    return redirect(url_for("po_list"))


# --- Helper for print/pdf ---
def get_purchaseorder_with_lines(id):
    db = get_db()
    po = db.execute("""
        SELECT po.*, s.name as supplier_name, s.address as supplier_address
        FROM purchase_orders po
        LEFT JOIN suppliers s ON po.supplier_id = s.id
        WHERE po.id = ?
    """, (id,)).fetchone()

    if not po:
        return None, None

    lines = db.execute("""
        SELECT l.*, i.sku, i.name as item_name, i.hsn_code
        FROM purchase_order_items l
        JOIN items i ON l.item_id = i.id
        WHERE l.purchase_order_id = ?
    """, (id,)).fetchall()

    return po, lines


# --- Print ---
@app.route("/purchase-orders/<int:id>/print", methods=["GET"], endpoint="po_print")
@login_required
def po_print(id):
    po, lines = get_purchaseorder_with_lines(id)
    if not po:
        flash("Purchase Order not found.", "danger")
        return redirect(url_for("po_list"))
    return render_template("print/purchase_order.html", po=po, lines=lines)


# --- PDF Export ---
@app.route("/purchase-orders/<int:id>/pdf", methods=["GET"], endpoint="po_pdf")
@login_required
def po_pdf(id):
    po, lines = get_purchaseorder_with_lines(id)
    if not po:
        flash("Purchase Order not found.", "danger")
        return redirect(url_for("po_list"))

    html = render_template("print/purchase_order.html", po=po, lines=lines)
    pdf_file = BytesIO()
    HTML(string=html, base_url=request.root_path).write_pdf(pdf_file)
    pdf_file.seek(0)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=f"PurchaseOrder_{id}.pdf",
        mimetype="application/pdf"
    )

# --- API Routes for suppliers (autocomplete / search) ---
@app.route("/api/suppliers/search")
@login_required
def search_suppliers():
    q = request.args.get("q", "").strip()
    limit = int(request.args.get("limit", 10))
    page = int(request.args.get("page", 1))
    offset = (page - 1) * limit

    db = get_db()
    rows = db.execute(
        """SELECT id, name, email, phone
           FROM suppliers
           WHERE name LIKE ? OR email LIKE ? OR phone LIKE ?
           ORDER BY name
           LIMIT ? OFFSET ?""",
        (f"%{q}%", f"%{q}%", f"%{q}%", limit, offset)
    ).fetchall()

    return jsonify({"results": [dict(r) for r in rows]})

# --- API Route for Items (Purchase Order autocomplete) ---
@app.route("/api/purchase-order/items/search")
@login_required
def po_search_items():
    q = request.args.get("q", "").strip()
    limit = int(request.args.get("limit", 10))
    page = int(request.args.get("page", 1))
    offset = (page - 1) * limit

    db = get_db()
    rows = db.execute(
        """
        SELECT
            id,
            name,
            sku,
            IFNULL(purchase_price, 0) AS purchase_price,
            IFNULL(gst_rate, 18) AS gst_rate,
            IFNULL(uom, 'Nos') AS uom,
            IFNULL(stock_qty, 0) AS stock_qty
        FROM items
        WHERE name LIKE ? COLLATE NOCASE
           OR sku LIKE ? COLLATE NOCASE
        ORDER BY name
        LIMIT ? OFFSET ?
        """,
        (f"%{q}%", f"%{q}%", limit, offset)
    ).fetchall()

    return jsonify({"results": [dict(r) for r in rows]})




# ----------------------
# Reports
# ----------------------

@app.route('/reports/outstanding')
@login_required
def reports_outstanding():
    db = get_db()
    # Customer outstanding
    cust_outstanding = db.execute('''
        SELECT c.id, c.name, c.mobile,
               SUM(si.grand_total) as total_invoiced,
               COALESCE(SUM(p.amount), 0) as total_paid,
               (SUM(si.grand_total) - COALESCE(SUM(p.amount), 0)) as balance
        FROM customers c
        JOIN sales_invoices si ON c.id = si.customer_id
        LEFT JOIN (SELECT invoice_id, SUM(amount) as amount FROM payments GROUP BY invoice_id) p ON si.id = p.invoice_id
        WHERE si.status = 'submitted'
        GROUP BY c.id
        HAVING balance > 0
        ORDER BY c.name
    ''').fetchall()

    # Supplier outstanding (payables)
    supp_outstanding = db.execute('''
        SELECT s.id, s.name, s.phone,
               SUM(pi.grand_total) as total_invoiced,
               COALESCE(SUM(p.amount), 0) as total_paid,
               (SUM(pi.grand_total) - COALESCE(SUM(p.amount), 0)) as balance
        FROM suppliers s
        JOIN purchase_invoices pi ON s.id = pi.supplier_id
        LEFT JOIN (SELECT purchase_invoice_id, SUM(amount) as amount FROM supplier_payments GROUP BY purchase_invoice_id) p ON pi.id = p.purchase_invoice_id
        WHERE pi.status = 'submitted'
        GROUP BY s.id
        HAVING balance > 0
        ORDER BY s.name
    ''').fetchall()
    
    return render_template('reports/outstanding.html', cust_outstanding=cust_outstanding, supp_outstanding=supp_outstanding)

@app.route('/reports/customer-ledger/')
@login_required
def report_customer_ledger():
    db = get_db()
    customers = db.execute("SELECT id, name FROM customers ORDER BY name").fetchall()
    
    customer_id = request.args.get('customer_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    ledger = []
    customer_name = None
    
    if customer_id:
        customer = db.execute("SELECT name FROM customers WHERE id = ?", (customer_id,)).fetchone()
        if customer:
            customer_name = customer['name']
            
            # Build query with date filters
            query = """
                SELECT 'invoice' as type, si.date, si.si_no as reference, si.grand_total as debit, 0 as credit
                FROM sales_invoices si
                WHERE si.customer_id = ? AND si.status = 'submitted'
                
                UNION ALL
                
                SELECT 'payment' as type, p.date, p.reference, 0 as debit, p.amount as credit
                FROM payments p
                JOIN sales_invoices si ON p.invoice_id = si.id
                WHERE si.customer_id = ?
            """
            params = [customer_id, customer_id]
            
            if start_date:
                query += " AND date(si.date) >= date(?)"
                params.append(start_date)
            if end_date:
                query += " AND date(si.date) <= date(?)"
                params.append(end_date)
                
            query += " ORDER BY date"
            
            ledger = db.execute(query, params).fetchall()

    return render_template('reports/customer_ledger.html', customers=customers, ledger=ledger, 
                           customer_id=customer_id, customer_name=customer_name,
                           start_date=start_date, end_date=end_date)

@app.route('/reports/supplier-ledger/')
@login_required
def report_supplier_ledger():
    db = get_db()
    suppliers = db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()
    
    supplier_id = request.args.get('supplier_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    ledger = []
    supplier_name = None
    
    if supplier_id:
        supplier = db.execute("SELECT name FROM suppliers WHERE id = ?", (supplier_id,)).fetchone()
        if supplier:
            supplier_name = supplier['name']
            
            query = """
                SELECT 'invoice' as type, pi.date, pi.pi_no as reference, pi.grand_total as debit, 0 as credit
                FROM purchase_invoices pi
                WHERE pi.supplier_id = ? AND pi.status = 'submitted'
                
                UNION ALL
                
                SELECT 'payment' as type, p.date, p.reference, 0 as debit, p.amount as credit
                FROM supplier_payments p
                JOIN purchase_invoices pi ON p.purchase_invoice_id = pi.id
                WHERE pi.supplier_id = ?
            """
            params = [supplier_id, supplier_id]
            
            if start_date:
                query += " AND date(pi.date) >= date(?)"
                params.append(start_date)
            if end_date:
                query += " AND date(pi.date) <= date(?)"
                params.append(end_date)
                
            query += " ORDER BY date"
            
            ledger = db.execute(query, params).fetchall()

    return render_template('reports/supplier_ledger.html', suppliers=suppliers, ledger=ledger, 
                           supplier_id=supplier_id, supplier_name=supplier_name,
                           start_date=start_date, end_date=end_date)

@app.route('/reports/stock-balance/')
@login_required
def report_stock_balance():
    db = get_db()
    items = db.execute("SELECT sku, name, stock_qty, uom, cost_price, selling_price FROM items ORDER BY name").fetchall()
    return render_template('reports/stock_balance.html', items=items)




@app.route('/print/supplier-ledger/<int:supplier_id>')
@login_required
def print_supplier_ledger(supplier_id):
    db = get_db()
    supplier = db.execute("SELECT * FROM suppliers WHERE id = ?", (supplier_id,)).fetchone()
    if not supplier:
        return "Supplier not found", 404

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = """
        SELECT 'invoice' as type, pi.date, pi.pi_no as reference, pi.grand_total as debit, 0 as credit
        FROM purchase_invoices pi
        WHERE pi.supplier_id = ? AND pi.status = 'submitted'
        UNION ALL
        SELECT 'payment' as type, p.date, p.reference, 0 as debit, p.amount as credit
        FROM supplier_payments p
        JOIN purchase_invoices pi ON p.purchase_invoice_id = pi.id
        WHERE pi.supplier_id = ?
    """
    params = [supplier_id, supplier_id]

    if start_date:
        query += " AND date(pi.date) >= date(?)"
        params.append(start_date)
    if end_date:
        query += " AND date(pi.date) <= date(?)"
        params.append(end_date)

    query += " ORDER BY date"
    ledger = db.execute(query, params).fetchall()

    rendered_html = render_template('print/supplier_ledger.html', supplier=supplier, ledger=ledger, start_date=start_date, end_date=end_date)
    return Response(
        HTML(string=rendered_html).write_pdf(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'inline; filename=supplier_ledger.pdf'}
    )


@app.route('/settings')
@login_required
def settings():
    return render_template('settings/index.html')

@app.route('/stock-moves')
@login_required
def stock_moves_list():
    db = get_db()
    moves = db.execute("""
        SELECT sm.id, sm.move_type, sm.qty, sm.reference, sm.notes, sm.created_at, i.name as item_name
        FROM stock_moves sm
        JOIN items i ON sm.item_id = i.id
        ORDER BY sm.created_at DESC
    """).fetchall()
    return render_template('stock_moves/list.html', moves=moves)

@app.route('/stock-moves/new', methods=['GET', 'POST'])
@login_required
def new_stock_move():
    db = get_db()
    if request.method == 'POST':
        item_id = request.form.get('item_id', type=int)
        move_type = request.form.get('move_type')
        qty = request.form.get('qty', type=float)
        reference = request.form.get('reference')
        notes = request.form.get('notes')

        if not all([item_id, move_type, qty]):
            flash('Missing required fields', 'danger')
            return redirect(url_for('new_stock_move'))

        db.execute("""
            INSERT INTO stock_moves (item_id, move_type, qty, reference, notes)
            VALUES (?, ?, ?, ?, ?)
        """, (item_id, move_type, qty, reference, notes))
        db.commit()

        flash('Stock move recorded', 'success')
        return redirect(url_for('stock_moves_list'))

    items = db.execute("SELECT id, name FROM items ORDER BY name").fetchall()
    return render_template('stock_moves/form.html', items=items)

@app.route('/delivery-notes')
@login_required
def delivery_notes_list():
    db = get_db()
    notes = db.execute("SELECT * FROM delivery_notes ORDER BY date DESC").fetchall()
    return render_template('delivery_notes/list.html', notes=notes)

@app.route('/delivery-notes/new', methods=['GET', 'POST'])
@login_required
def new_delivery_note():
    db = get_db()
    if request.method == 'POST':
        # ... (implementation for creating a new delivery note)
        pass
    customers = db.execute("SELECT id, name FROM customers ORDER BY name").fetchall()
    return render_template('delivery_notes/new.html', customers=customers)

@app.route('/delivery-notes/<int:id>')
@login_required
def delivery_note_detail(id):
    db = get_db()
    note = db.execute("SELECT * FROM delivery_notes WHERE id = ?", (id,)).fetchone()
    items = db.execute("SELECT * FROM delivery_note_items WHERE delivery_note_id = ?", (id,)).fetchall()
    return render_template('delivery_notes/detail.html', note=note, items=items)

@app.route('/purchase-returns')
@login_required
def purchase_returns_list():
    db = get_db()
    returns = db.execute("SELECT * FROM purchase_returns ORDER BY date DESC").fetchall()
    return render_template('purchase_returns/list.html', returns=returns)

@app.route('/purchase-returns/new', methods=['GET', 'POST'])
@login_required
def new_purchase_return():
    db = get_db()
    if request.method == 'POST':
        # ... (implementation for creating a new purchase return)
        pass
    suppliers = db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()
    return render_template('purchase_returns/new.html', suppliers=suppliers)

@app.route('/purchase-returns/<int:id>')
@login_required
def purchase_return_detail(id):
    db = get_db()
    preturn = db.execute("SELECT * FROM purchase_returns WHERE id = ?", (id,)).fetchone()
    items = db.execute("SELECT * FROM purchase_return_items WHERE purchase_return_id = ?", (id,)).fetchall()
    return render_template('purchase_returns/detail.html', preturn=preturn, items=items)

@app.route('/purchase-returns/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_purchase_return(id):
    db = get_db()
    preturn = db.execute("SELECT * FROM purchase_returns WHERE id = ?", (id,)).fetchone()
    if request.method == 'POST':
        # ... (implementation for editing a purchase return)
        pass
    return render_template('purchase_returns/edit.html', preturn=preturn)

@app.route('/print/delivery-note/<int:id>')
@login_required
def print_delivery_note(id):
    db = get_db()
    note = db.execute("SELECT * FROM delivery_notes WHERE id = ?", (id,)).fetchone()
    items = db.execute("SELECT * FROM delivery_note_items WHERE delivery_note_id = ?", (id,)).fetchall()
    rendered_html = render_template('print/delivery_note.html', note=note, items=items)
    return Response(
        HTML(string=rendered_html).write_pdf(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'inline; filename=delivery_note.pdf'}
    )

@app.route('/print/packing-slip/<int:id>')
@login_required
def print_packing_slip(id):
    db = get_db()
    note = db.execute("SELECT * FROM delivery_notes WHERE id = ?", (id,)).fetchone()
    items = db.execute("SELECT * FROM delivery_note_items WHERE delivery_note_id = ?", (id,)).fetchall()
    rendered_html = render_template('print/packing_slip.html', note=note, items=items)
    return Response(
        HTML(string=rendered_html).write_pdf(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'inline; filename=packing_slip.pdf'}
    )

@app.route('/print/purchase-return/<int:id>')
@login_required
def print_purchase_return(id):
    db = get_db()
    preturn = db.execute("SELECT * FROM purchase_returns WHERE id = ?", (id,)).fetchone()
    items = db.execute("SELECT * FROM purchase_return_items WHERE purchase_return_id = ?", (id,)).fetchall()
    rendered_html = render_template('print/purchase_return.html', preturn=preturn, items=items)
    return Response(
        HTML(string=rendered_html).write_pdf(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'inline; filename=purchase_return.pdf'}
    )

@app.route('/print/purchase-return-delivery-note/<int:id>')
@login_required
def print_purchase_return_delivery_note(id):
    db = get_db()
    preturn = db.execute("SELECT * FROM purchase_returns WHERE id = ?", (id,)).fetchone()
    items = db.execute("SELECT * FROM purchase_return_items WHERE purchase_return_id = ?", (id,)).fetchall()
    rendered_html = render_template('print/purchase_return_delivery_note.html', preturn=preturn, items=items)
    return Response(
        HTML(string=rendered_html).write_pdf(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'inline; filename=purchase_return_delivery_note.pdf'}
    )

@app.route('/print/purchase-return-packing-slip/<int:id>')
@login_required
def print_purchase_return_packing_slip(id):
    db = get_db()
    preturn = db.execute("SELECT * FROM purchase_returns WHERE id = ?", (id,)).fetchone()
    items = db.execute("SELECT * FROM purchase_return_items WHERE purchase_return_id = ?", (id,)).fetchall()
    rendered_html = render_template('print/purchase_return_packing_slip.html', preturn=preturn, items=items)
    return Response(
        HTML(string=rendered_html).write_pdf(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'inline; filename=purchase_return_packing_slip.pdf'}
    )


@app.route('/reports/outstanding')
@login_required
def report_outstanding():
    db = get_db()
    # ... (implementation for outstanding report)
    return render_template('reports/outstanding.html')

@app.route('/reports/supplier-outstanding')
@login_required
def report_supplier_outstanding():
    db = get_db()
    # ... (implementation for supplier outstanding report)
    return render_template('reports/supplier_outstanding.html')

@app.route('/sales-invoices/<int:id>/add-payment', methods=['GET', 'POST'])
@login_required
def add_payment_to_invoice(id):
    db = get_db()
    invoice = db.execute("SELECT * FROM sales_invoices WHERE id = ?", (id,)).fetchone()
    if request.method == 'POST':
        # ... (implementation for adding payment)
        pass
    return render_template('sales_invoices/add_payment.html', invoice=invoice)

@app.route('/purchase-invoices/<int:id>/add-payment', methods=['GET', 'POST'])
@login_required
def add_payment_to_purchase_invoice(id):
    db = get_db()
    invoice = db.execute("SELECT * FROM purchase_invoices WHERE id = ?", (id,)).fetchone()
    if request.method == 'POST':
        # ... (implementation for adding payment)
        pass
    return render_template('purchase_invoices/add_payment.html', invoice=invoice)
    
@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))
   


if __name__ == '__main__':
    with app.app_context():
        init_db()
    app.run(debug=True, port=5010)